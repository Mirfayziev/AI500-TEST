from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from functools import wraps
import os
import io
import json
from sqlalchemy import func, or_, and_

# Import models and config
from models import *
from config import Config

# Import for Excel/PDF generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(Config)

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Iltimos, tizimga kiring'

# Create upload folders
UPLOAD_FOLDERS = ['uploads', 'uploads/vehicles', 'uploads/buildings', 'uploads/users',
                  'uploads/tasks', 'uploads/contracts', 'uploads/outsourcing', 
                  'uploads/guests', 'uploads/greenspaces', 'uploads/solarpanels']

for folder in UPLOAD_FOLDERS:
    os.makedirs(folder, exist_ok=True)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Custom decorators
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Bu sahifaga kirish uchun admin huquqi kerak', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def rahbar_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['admin', 'rahbar']:
            flash('Bu sahifaga kirish uchun rahbar huquqi kerak', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def module_access_required(module_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not current_user.has_module_access(module_name):
                flash(f'{module_name} moduliga kirish huquqingiz yo\'q', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Helper functions
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def save_file(file, folder):
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
        filename = timestamp + filename
        filepath = os.path.join(folder, filename)
        file.save(filepath)
        return filepath
    return None

def get_task_status_color(task):
    """Returns color indicator for task status"""
    if task.status == 'completed':
        return 'success'
    if task.due_date:
        days_left = (task.due_date - datetime.utcnow()).days
        if days_left < 0:
            return 'danger'  # red - overdue
        elif days_left <= 3:
            return 'danger'  # red - 3 days or less
        elif days_left <= 5:
            return 'warning'  # orange - 5 days or less
        elif days_left <= 7:
            return 'info'  # yellow - 7 days or less
    return 'primary'

def send_telegram_notification(user_id, message):
    """Send notification via Telegram bot"""
    try:
        user = User.query.get(user_id)
        if user and user.telegram_chat_id and app.config['TELEGRAM_BOT_TOKEN']:
            import requests
            url = f"https://api.telegram.org/bot{app.config['TELEGRAM_BOT_TOKEN']}/sendMessage"
            data = {
                'chat_id': user.telegram_chat_id,
                'text': message,
                'parse_mode': 'HTML'
            }
            requests.post(url, data=data)
    except Exception as e:
        print(f"Telegram notification error: {e}")

# ==================== AUTHENTICATION ROUTES ====================

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))



@app.route('/demo')
def demo():
    """Public AI500 competition demo landing page"""
    return render_template('demo.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = request.form.get('remember', False)
        
        user = User.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            if not user.is_active:
                flash('Sizning hisobingiz faol emas', 'danger')
                return redirect(url_for('login'))
            
            login_user(user, remember=remember)
            
            # Log activity
            log = ActivityLog(
                user_id=user.id,
                action='Login',
                module='auth',
                details='User logged in successfully',
                ip_address=request.remote_addr
            )
            db.session.add(log)
            db.session.commit()
            
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Email yoki parol noto\'g\'ri', 'danger')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        password = request.form.get('password')
        telegram_username = request.form.get('telegram_username')
        department = request.form.get('department')
        position = request.form.get('position')
        bio = request.form.get('bio')
        
        if User.query.filter_by(email=email).first():
            flash('Bu email allaqachon ro\'yxatdan o\'tgan', 'danger')
            return redirect(url_for('register'))
        
        user = User(
            full_name=full_name,
            email=email,
            telegram_username=telegram_username,
            department=department,
            position=position,
            bio=bio,
            role='user'
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Ro\'yxatdan o\'tish muvaffaqiyatli! Iltimos, tizimga kiring', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action='Logout',
        module='auth',
        details='User logged out',
        ip_address=request.remote_addr
    )
    db.session.add(log)
    db.session.commit()
    
    logout_user()
    flash('Tizimdan chiqildi', 'success')
    return redirect(url_for('login'))

# ==================== DASHBOARD ====================

@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics for dashboard
    stats = {}
    
    # Tasks statistics
    if current_user.role in ['admin', 'rahbar']:
        stats['total_tasks'] = Task.query.count()
        stats['pending_tasks'] = Task.query.filter_by(status='pending').count()
        stats['in_progress_tasks'] = Task.query.filter_by(status='in_progress').count()
        stats['completed_tasks'] = Task.query.filter_by(status='completed').count()
        stats['overdue_tasks'] = Task.query.filter(
            Task.due_date < datetime.utcnow(),
            Task.status != 'completed'
        ).count()
    else:
        user_task_ids = [ta.task_id for ta in current_user.tasks_assigned]
        stats['total_tasks'] = len(user_task_ids)
        stats['pending_tasks'] = Task.query.filter(
            Task.id.in_(user_task_ids),
            Task.status == 'pending'
        ).count()
        stats['in_progress_tasks'] = Task.query.filter(
            Task.id.in_(user_task_ids),
            Task.status == 'in_progress'
        ).count()
        stats['completed_tasks'] = Task.query.filter(
            Task.id.in_(user_task_ids),
            Task.status == 'completed'
        ).count()
        stats['overdue_tasks'] = Task.query.filter(
            Task.id.in_(user_task_ids),
            Task.due_date < datetime.utcnow(),
            Task.status != 'completed'
        ).count()
    
    # Other statistics
    stats['total_vehicles'] = Vehicle.query.count()
    stats['active_vehicles'] = Vehicle.query.filter_by(status='active').count()
    stats['total_buildings'] = Building.query.count()
    stats['total_employees'] = User.query.filter_by(role='xodim').count()
    stats['total_contracts'] = Contract.query.count()
    stats['active_contracts'] = Contract.query.filter_by(status='active').count()
    stats['total_guests'] = Guest.query.count()
    stats['total_organizations'] = Organization.query.count()
    
    # Recent activities
    recent_tasks = []
    if current_user.role in ['admin', 'rahbar']:
        recent_tasks = Task.query.order_by(Task.created_at.desc()).limit(5).all()
    else:
        user_task_ids = [ta.task_id for ta in current_user.tasks_assigned]
        recent_tasks = Task.query.filter(Task.id.in_(user_task_ids)).order_by(
            Task.created_at.desc()
        ).limit(5).all()
    
    # Get user's assigned modules
    user_modules = [um.module_name for um in current_user.assigned_modules]
    
    # Notifications
    notifications = Notification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).order_by(Notification.created_at.desc()).limit(10).all()
    
    return render_template('dashboard.html',
                         stats=stats,
                         recent_tasks=recent_tasks,
                         user_modules=user_modules,
                         notifications=notifications,
                         get_task_status_color=get_task_status_color)


# ==================== TASKS MODULE ====================

@app.route('/tasks')
@login_required
@module_access_required('tasks')
def tasks():
    if current_user.role in ['admin', 'rahbar']:
        tasks = Task.query.order_by(Task.created_at.desc()).all()
    else:
        user_task_ids = [ta.task_id for ta in current_user.tasks_assigned]
        tasks = Task.query.filter(Task.id.in_(user_task_ids)).order_by(Task.created_at.desc()).all()
    
    return render_template('tasks/index.html', tasks=tasks, get_task_status_color=get_task_status_color)

@app.route('/tasks/create', methods=['GET', 'POST'])
@login_required
@admin_required
def tasks_create():
    if request.method == 'POST':
        task = Task(
            title=request.form.get('title'),
            description=request.form.get('description'),
            priority=request.form.get('priority', 'medium'),
            status='pending',
            start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d') if request.form.get('start_date') else None,
            due_date=datetime.strptime(request.form.get('due_date'), '%Y-%m-%d') if request.form.get('due_date') else None,
            created_by=current_user.id
        )
        db.session.add(task)
        db.session.commit()
        
        # Assign users
        assigned_user_ids = request.form.getlist('assigned_users')
        for user_id in assigned_user_ids:
            assignment = TaskAssignment(
                task_id=task.id,
                user_id=int(user_id),
                assigned_by=current_user.id
            )
            db.session.add(assignment)
            
            # Send notification
            notification = Notification(
                user_id=int(user_id),
                title='Yangi topshiriq',
                message=f'Sizga yangi topshiriq biriktirildi: {task.title}',
                type='task',
                link=url_for('tasks_view', id=task.id)
            )
            db.session.add(notification)
            
            # Send Telegram notification
            send_telegram_notification(
                int(user_id),
                f"📋 <b>Yangi topshiriq</b>\n\n"
                f"<b>Nomi:</b> {task.title}\n"
                f"<b>Muhimlik:</b> {task.priority}\n"
                f"<b>Muddat:</b> {task.due_date.strftime('%d.%m.%Y') if task.due_date else 'Belgilanmagan'}"
            )
        
        db.session.commit()
        
        flash('Topshiriq muvaffaqiyatli yaratildi', 'success')
        return redirect(url_for('tasks'))
    
    users = User.query.filter_by(role='xodim', is_active=True).all()
    return render_template('tasks/create.html', users=users)

@app.route('/tasks/<int:id>')
@login_required
@module_access_required('tasks')
def tasks_view(id):
    task = Task.query.get_or_404(id)
    
    # Check access - XODIMLAR FAQAT O'Z TOPSHIRIQLARINI KO'RADI
    if current_user.role == 'xodim':
        # Tekshirish - bu topshiriq xodimga biriktirilganmi?
        is_assigned = any(ta.user_id == current_user.id for ta in task.assignments)
        if not is_assigned:
            flash('Bu topshiriqni ko\'rish huquqingiz yo\'q', 'danger')
            return redirect(url_for('tasks'))
    elif current_user.role not in ['admin', 'rahbar']:
        # User role uchun ham cheklash
        is_assigned = any(ta.user_id == current_user.id for ta in task.assignments)
        if not is_assigned:
            flash('Bu topshiriqni ko\'rish huquqingiz yo\'q', 'danger')
            return redirect(url_for('tasks'))
    
    comments = TaskComment.query.filter_by(task_id=id).order_by(TaskComment.created_at.desc()).all()
    
    return render_template('tasks/view.html', 
                         task=task, 
                         comments=comments,
                         get_task_status_color=get_task_status_color)

@app.route('/tasks/<int:id>/update-status', methods=['POST'])
@login_required
def tasks_update_status(id):
    task = Task.query.get_or_404(id)
    new_status = request.form.get('status')
    
    if new_status == 'completed':
        # Xodim tugalladi deb belgilasa, rahbar tasdiqini kutish
        if current_user.role == 'xodim':
            task.status = 'review'
            flash('Topshiriq rahbar tasdiqini kutmoqda', 'info')
            
            # Notify rahbar
            notification = Notification(
                user_id=task.created_by,
                title='Topshiriq tasdiqlash kutilmoqda',
                message=f'{current_user.full_name} "{task.title}" topshiriqni tugatdi',
                type='task',
                link=url_for('tasks_view', id=task.id)
            )
            db.session.add(notification)
        else:
            # Rahbar tasdiqlaganda
            task.status = 'completed'
            task.completion_date = datetime.utcnow()
            flash('Topshiriq tasdiqlandi va bajarildi deb belgilandi', 'success')
            
            # Notify assigned users
            for assignment in task.assignments:
                notification = Notification(
                    user_id=assignment.user_id,
                    title='Topshiriq tasdiqlandi',
                    message=f'"{task.title}" topshiriqi tasdiqlandi',
                    type='task',
                    link=url_for('tasks_view', id=task.id)
                )
                db.session.add(notification)
    else:
        task.status = new_status
    
    task.updated_at = datetime.utcnow()
    db.session.commit()
    
    return redirect(url_for('tasks_view', id=id))

@app.route('/tasks/<int:id>/add-comment', methods=['POST'])
@login_required
def tasks_add_comment(id):
    task = Task.query.get_or_404(id)
    comment_text = request.form.get('comment')
    
    if comment_text:
        comment = TaskComment(
            task_id=id,
            user_id=current_user.id,
            comment=comment_text
        )
        db.session.add(comment)
        db.session.commit()
        
        flash('Izoh qo\'shildi', 'success')
    
    return redirect(url_for('tasks_view', id=id))

@app.route('/tasks/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def tasks_delete(id):
    task = Task.query.get_or_404(id)
    db.session.delete(task)
    db.session.commit()
    
    flash('Topshiriq o\'chirildi', 'success')
    return redirect(url_for('tasks'))

# ==================== VEHICLES MODULE ====================

@app.route('/vehicles')
@login_required
@module_access_required('vehicles')
def vehicles():
    vehicles = Vehicle.query.order_by(Vehicle.created_at.desc()).all()
    return render_template('vehicles/index.html', vehicles=vehicles)

@app.route('/vehicles/create', methods=['GET', 'POST'])
@login_required
@module_access_required('vehicles')
def vehicles_create():
    if request.method == 'POST':
        vehicle = Vehicle(
            brand=request.form.get('brand'),
            model=request.form.get('model'),
            year=int(request.form.get('year')) if request.form.get('year') else None,
            license_plate=request.form.get('license_plate'),
            vin_number=request.form.get('vin_number'),
            color=request.form.get('color'),
            status=request.form.get('status', 'active'),
            last_maintenance=datetime.strptime(request.form.get('last_maintenance'), '%Y-%m-%d') if request.form.get('last_maintenance') else None,
            next_maintenance=datetime.strptime(request.form.get('next_maintenance'), '%Y-%m-%d') if request.form.get('next_maintenance') else None,
            driver_id=int(request.form.get('driver_id')) if request.form.get('driver_id') else None,
            defects=request.form.get('defects'),
            notes=request.form.get('notes')
        )
        
        # Handle photo upload
        if 'photo' in request.files:
            file = request.files['photo']
            filepath = save_file(file, 'uploads/vehicles')
            if filepath:
                vehicle.photo = filepath
        
        db.session.add(vehicle)
        db.session.commit()
        
        flash('Transport muvaffaqiyatli qo\'shildi', 'success')
        return redirect(url_for('vehicles'))
    
    drivers = User.query.filter_by(role='xodim', is_active=True).all()
    return render_template('vehicles/create.html', drivers=drivers)

@app.route('/vehicles/<int:id>')
@login_required
@module_access_required('vehicles')
def vehicles_view(id):
    vehicle = Vehicle.query.get_or_404(id)
    return render_template('vehicles/view.html', vehicle=vehicle)

@app.route('/vehicles/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@module_access_required('vehicles')
def vehicles_edit(id):
    vehicle = Vehicle.query.get_or_404(id)
    
    if request.method == 'POST':
        vehicle.brand = request.form.get('brand')
        vehicle.model = request.form.get('model')
        vehicle.year = int(request.form.get('year')) if request.form.get('year') else None
        vehicle.license_plate = request.form.get('license_plate')
        vehicle.vin_number = request.form.get('vin_number')
        vehicle.color = request.form.get('color')
        vehicle.status = request.form.get('status')
        vehicle.last_maintenance = datetime.strptime(request.form.get('last_maintenance'), '%Y-%m-%d') if request.form.get('last_maintenance') else None
        vehicle.next_maintenance = datetime.strptime(request.form.get('next_maintenance'), '%Y-%m-%d') if request.form.get('next_maintenance') else None
        vehicle.driver_id = int(request.form.get('driver_id')) if request.form.get('driver_id') else None
        vehicle.defects = request.form.get('defects')
        vehicle.notes = request.form.get('notes')
        
        if 'photo' in request.files:
            file = request.files['photo']
            filepath = save_file(file, 'uploads/vehicles')
            if filepath:
                vehicle.photo = filepath
        
        db.session.commit()
        
        flash('Transport ma\'lumotlari yangilandi', 'success')
        return redirect(url_for('vehicles_view', id=id))
    
    drivers = User.query.filter_by(role='xodim', is_active=True).all()
    return render_template('vehicles/edit.html', vehicle=vehicle, drivers=drivers)

@app.route('/vehicles/<int:id>/delete', methods=['POST'])
@login_required
@rahbar_required
def vehicles_delete(id):
    vehicle = Vehicle.query.get_or_404(id)
    db.session.delete(vehicle)
    db.session.commit()
    
    flash('Transport o\'chirildi', 'success')
    return redirect(url_for('vehicles'))

@app.route('/vehicles/export-pdf')
@login_required
@module_access_required('vehicles')
def vehicles_export_pdf():
    vehicles = Vehicle.query.all()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1
    )
    title = Paragraph("Transport Vositalari Hisoboti", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    data = [['№', 'Markasi', 'Modeli', 'Raqami', 'Yili', 'Holati', 'Haydovchi']]
    
    for idx, vehicle in enumerate(vehicles, 1):
        driver_name = vehicle.driver.full_name if vehicle.driver else '-'
        data.append([
            str(idx),
            vehicle.brand or '-',
            vehicle.model or '-',
            vehicle.license_plate or '-',
            str(vehicle.year) if vehicle.year else '-',
            vehicle.status or '-',
            driver_name
        ])
    
    # Create table
    table = Table(data, colWidths=[0.5*inch, 1.2*inch, 1.2*inch, 1*inch, 0.8*inch, 1*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'transport_hisoboti_{datetime.now().strftime("%Y%m%d")}.pdf',
        mimetype='application/pdf'
    )

@app.route('/vehicles/export-excel')
@login_required
@module_access_required('vehicles')
def vehicles_export_excel():
    vehicles = Vehicle.query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transport Vositalari"
    
    # Headers
    headers = ['№', 'Markasi', 'Modeli', 'Raqami', 'Yili', 'Rangi', 'Holati', 'Haydovchi', 
               'So\'nggi remont', 'Keyingi remont', 'Defektlar']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for idx, vehicle in enumerate(vehicles, 1):
        driver_name = vehicle.driver.full_name if vehicle.driver else '-'
        last_maintenance = vehicle.last_maintenance.strftime('%d.%m.%Y') if vehicle.last_maintenance else '-'
        next_maintenance = vehicle.next_maintenance.strftime('%d.%m.%Y') if vehicle.next_maintenance else '-'
        
        ws.append([
            idx,
            vehicle.brand or '-',
            vehicle.model or '-',
            vehicle.license_plate or '-',
            vehicle.year or '-',
            vehicle.color or '-',
            vehicle.status or '-',
            driver_name,
            last_maintenance,
            next_maintenance,
            vehicle.defects or '-'
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'transport_hisoboti_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== BUILDINGS MODULE ====================

@app.route('/buildings')
@login_required
@module_access_required('buildings')
def buildings():
    buildings = Building.query.order_by(Building.created_at.desc()).all()
    categories = BuildingCategory.query.all()
    return render_template('buildings/index.html', buildings=buildings, categories=categories)

@app.route('/buildings/categories', methods=['GET', 'POST'])
@login_required
@module_access_required('buildings')
def buildings_categories():
    if request.method == 'POST':
        category = BuildingCategory(
            name=request.form.get('name'),
            description=request.form.get('description'),
            icon=request.form.get('icon'),
            created_by=current_user.id
        )
        db.session.add(category)
        db.session.commit()
        
        flash('Kategoriya qo\'shildi', 'success')
        return redirect(url_for('buildings_categories'))
    
    categories = BuildingCategory.query.order_by(BuildingCategory.created_at.desc()).all()
    return render_template('buildings/categories.html', categories=categories)

@app.route('/buildings/create', methods=['GET', 'POST'])
@login_required
@module_access_required('buildings')
def buildings_create():
    if request.method == 'POST':
        building = Building(
            category_id=int(request.form.get('category_id')) if request.form.get('category_id') else None,
            name=request.form.get('name'),
            address=request.form.get('address'),
            area=float(request.form.get('area')) if request.form.get('area') else None,
            floors=int(request.form.get('floors')) if request.form.get('floors') else None,
            rooms=int(request.form.get('rooms')) if request.form.get('rooms') else None,
            construction_year=int(request.form.get('construction_year')) if request.form.get('construction_year') else None,
            status=request.form.get('status'),
            description=request.form.get('description'),
            created_by=current_user.id
        )
        
        if 'photo' in request.files:
            file = request.files['photo']
            filepath = save_file(file, 'uploads/buildings')
            if filepath:
                building.photo = filepath
        
        db.session.add(building)
        db.session.commit()
        
        flash('Bino muvaffaqiyatli qo\'shildi', 'success')
        return redirect(url_for('buildings'))
    
    categories = BuildingCategory.query.all()
    return render_template('buildings/create.html', categories=categories)

# ==================== GREEN SPACES MODULE ====================

@app.route('/greenspaces')
@login_required
@module_access_required('greenspaces')
def greenspaces():
    greenspaces = GreenSpace.query.order_by(GreenSpace.created_at.desc()).all()
    categories = GreenSpaceCategory.query.all()
    return render_template('greenspaces/index.html', greenspaces=greenspaces, categories=categories)

@app.route('/greenspaces/categories', methods=['GET', 'POST'])
@login_required
@module_access_required('greenspaces')
def greenspaces_categories():
    if request.method == 'POST':
        category = GreenSpaceCategory(
            name=request.form.get('name'),
            description=request.form.get('description'),
            icon=request.form.get('icon'),
            created_by=current_user.id
        )
        db.session.add(category)
        db.session.commit()
        
        flash('Kategoriya qo\'shildi', 'success')
        return redirect(url_for('greenspaces_categories'))
    
    categories = GreenSpaceCategory.query.order_by(GreenSpaceCategory.created_at.desc()).all()
    return render_template('greenspaces/categories.html', categories=categories)

@app.route('/greenspaces/create', methods=['GET', 'POST'])
@login_required
@module_access_required('greenspaces')
def greenspaces_create():
    if request.method == 'POST':
        greenspace = GreenSpace(
            category_id=int(request.form.get('category_id')) if request.form.get('category_id') else None,
            name=request.form.get('name'),
            location=request.form.get('location'),
            area=float(request.form.get('area')) if request.form.get('area') else None,
            plant_types=request.form.get('plant_types'),
            maintenance_schedule=request.form.get('maintenance_schedule'),
            status=request.form.get('status'),
            description=request.form.get('description'),
            created_by=current_user.id
        )
        
        if 'photo' in request.files:
            file = request.files['photo']
            filepath = save_file(file, 'uploads/greenspaces')
            if filepath:
                greenspace.photo = filepath
        
        db.session.add(greenspace)
        db.session.commit()
        
        flash('Yashil makon muvaffaqiyatli qo\'shildi', 'success')
        return redirect(url_for('greenspaces'))
    
    categories = GreenSpaceCategory.query.all()
    return render_template('greenspaces/create.html', categories=categories)

# ==================== SOLAR PANELS MODULE ====================

@app.route('/solarpanels')
@login_required
@module_access_required('solarpanels')
def solarpanels():
    solarpanels = SolarPanel.query.order_by(SolarPanel.created_at.desc()).all()
    return render_template('solarpanels/index.html', solarpanels=solarpanels)

@app.route('/solarpanels/create', methods=['GET', 'POST'])
@login_required
@module_access_required('solarpanels')
def solarpanels_create():
    if request.method == 'POST':
        panel = SolarPanel(
            building_id=int(request.form.get('building_id')) if request.form.get('building_id') else None,
            panel_type=request.form.get('panel_type'),
            capacity=float(request.form.get('capacity')) if request.form.get('capacity') else None,
            installation_date=datetime.strptime(request.form.get('installation_date'), '%Y-%m-%d') if request.form.get('installation_date') else None,
            manufacturer=request.form.get('manufacturer'),
            model=request.form.get('model'),
            efficiency=float(request.form.get('efficiency')) if request.form.get('efficiency') else None,
            status=request.form.get('status'),
            monitoring_url=request.form.get('monitoring_url'),
            notes=request.form.get('notes')
        )
        
        if 'photo' in request.files:
            file = request.files['photo']
            filepath = save_file(file, 'uploads/solarpanels')
            if filepath:
                panel.photo = filepath
        
        db.session.add(panel)
        db.session.commit()
        
        flash('Quyosh paneli muvaffaqiyatli qo\'shildi', 'success')
        return redirect(url_for('solarpanels'))
    
    buildings = Building.query.all()
    return render_template('solarpanels/create.html', buildings=buildings)

# ==================== EMPLOYEES MODULE ====================

@app.route('/employees')
@login_required
@rahbar_required
def employees():
    employees = User.query.filter_by(role='xodim').order_by(User.created_at.desc()).all()
    return render_template('employees/index.html', employees=employees)

@app.route('/employees/<int:id>')
@login_required
def employees_view(id):
    employee = User.query.get_or_404(id)
    
    # Get employee statistics
    if employee.role == 'xodim':
        assigned_tasks = [ta.task for ta in employee.tasks_assigned]
        completed_tasks = [t for t in assigned_tasks if t.status == 'completed']
        pending_tasks = [t for t in assigned_tasks if t.status in ['pending', 'in_progress']]
        
        stats = {
            'total_tasks': len(assigned_tasks),
            'completed_tasks': len(completed_tasks),
            'pending_tasks': len(pending_tasks),
            'completion_rate': round((len(completed_tasks) / len(assigned_tasks) * 100) if assigned_tasks else 0, 1)
        }
    else:
        stats = None
    
    return render_template('employees/view.html', employee=employee, stats=stats)

@app.route('/employees/<int:id>/assign-module', methods=['POST'])
@login_required
@rahbar_required
def employees_assign_module(id):
    employee = User.query.get_or_404(id)
    module_name = request.form.get('module_name')
    
    # Check if already assigned
    existing = UserModule.query.filter_by(user_id=id, module_name=module_name).first()
    if existing:
        flash('Bu modul allaqachon biriktirilgan', 'warning')
    else:
        assignment = UserModule(
            user_id=id,
            module_name=module_name,
            assigned_by=current_user.id
        )
        db.session.add(assignment)
        db.session.commit()
        
        flash('Modul muvaffaqiyatli biriktirildi', 'success')
    
    return redirect(url_for('employees_view', id=id))

@app.route('/employees/<int:id>/remove-module/<module_name>', methods=['POST'])
@login_required
@rahbar_required
def employees_remove_module(id, module_name):
    assignment = UserModule.query.filter_by(user_id=id, module_name=module_name).first()
    if assignment:
        db.session.delete(assignment)
        db.session.commit()
        flash('Modul olib tashlandi', 'success')
    
    return redirect(url_for('employees_view', id=id))

# ==================== OUTSOURCING MODULE ====================

@app.route('/outsourcing')
@login_required
@module_access_required('outsourcing')
def outsourcing():
    services = OutsourcingService.query.order_by(OutsourcingService.created_at.desc()).all()
    return render_template('outsourcing/index.html', services=services)

@app.route('/outsourcing/create', methods=['GET', 'POST'])
@login_required
@module_access_required('outsourcing')
def outsourcing_create():
    if request.method == 'POST':
        service = OutsourcingService(
            service_name=request.form.get('service_name'),
            provider_name=request.form.get('provider_name'),
            contract_number=request.form.get('contract_number'),
            contract_date=datetime.strptime(request.form.get('contract_date'), '%Y-%m-%d') if request.form.get('contract_date') else None,
            start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d') if request.form.get('start_date') else None,
            end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d') if request.form.get('end_date') else None,
            service_description=request.form.get('service_description'),
            cost=float(request.form.get('cost')) if request.form.get('cost') else None,
            status=request.form.get('status'),
            contact_person=request.form.get('contact_person'),
            contact_phone=request.form.get('contact_phone'),
            notes=request.form.get('notes'),
            created_by=current_user.id
        )
        
        db.session.add(service)
        db.session.commit()
        
        flash('Outsorsing xizmati qo\'shildi', 'success')
        return redirect(url_for('outsourcing'))
    
    return render_template('outsourcing/create.html')

# ==================== ORGANIZATIONS MODULE ====================

@app.route('/organizations')
@login_required
@module_access_required('organizations')
def organizations():
    organizations = Organization.query.order_by(Organization.created_at.desc()).all()
    return render_template('organizations/index.html', organizations=organizations)

@app.route('/organizations/create', methods=['GET', 'POST'])
@login_required
@module_access_required('organizations')
def organizations_create():
    if request.method == 'POST':
        org = Organization(
            name=request.form.get('name'),
            employee_count=int(request.form.get('employee_count')) if request.form.get('employee_count') else None,
            building_area=float(request.form.get('building_area')) if request.form.get('building_area') else None,
            description=request.form.get('description'),
            address=request.form.get('address'),
            phone=request.form.get('phone'),
            email=request.form.get('email'),
            website=request.form.get('website'),
            vehicles=request.form.get('vehicles'),
            established_date=datetime.strptime(request.form.get('established_date'), '%Y-%m-%d') if request.form.get('established_date') else None,
            created_by=current_user.id
        )
        
        if 'logo' in request.files:
            file = request.files['logo']
            filepath = save_file(file, 'uploads')
            if filepath:
                org.logo = filepath
        
        db.session.add(org)
        db.session.commit()
        
        flash('Tashkilot qo\'shildi', 'success')
        return redirect(url_for('organizations'))
    
    return render_template('organizations/create.html')

# ==================== GUESTS MODULE ====================

@app.route('/guests')
@login_required
@module_access_required('guests')
def guests():
    guests = Guest.query.order_by(Guest.created_at.desc()).all()
    return render_template('guests/index.html', guests=guests)

@app.route('/guests/create', methods=['GET', 'POST'])
@login_required
@module_access_required('guests')
def guests_create():
    if request.method == 'POST':
        guest = Guest(
            full_name=request.form.get('full_name'),
            organization=request.form.get('organization'),
            position=request.form.get('position'),
            arrival_date=datetime.strptime(request.form.get('arrival_date'), '%Y-%m-%d') if request.form.get('arrival_date') else None,
            departure_date=datetime.strptime(request.form.get('departure_date'), '%Y-%m-%d') if request.form.get('departure_date') else None,
            visit_purpose=request.form.get('visit_purpose'),
            reference_number=request.form.get('reference_number'),
            services_provided=request.form.get('services_provided'),
            restaurant_expense=float(request.form.get('restaurant_expense')) if request.form.get('restaurant_expense') else 0,
            gift_expense=float(request.form.get('gift_expense')) if request.form.get('gift_expense') else 0,
            other_expenses=float(request.form.get('other_expenses')) if request.form.get('other_expenses') else 0,
            notes=request.form.get('notes'),
            created_by=current_user.id
        )
        
        # Calculate total
        guest.total_expense = guest.restaurant_expense + guest.gift_expense + guest.other_expenses
        
        if 'photo' in request.files:
            file = request.files['photo']
            filepath = save_file(file, 'uploads/guests')
            if filepath:
                guest.photo = filepath
        
        db.session.add(guest)
        db.session.commit()
        
        flash('Mehmon ma\'lumotlari qo\'shildi', 'success')
        return redirect(url_for('guests'))
    
    return render_template('guests/create.html')

@app.route('/guests/<int:id>/export-pdf')
@login_required
@module_access_required('guests')
def guests_export_pdf(id):
    guest = Guest.query.get_or_404(id)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph("Mehmon Ma'lumotlari", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Guest info table
    data = [
        ['To\'liq ismi:', guest.full_name or '-'],
        ['Tashkilot:', guest.organization or '-'],
        ['Lavozim:', guest.position or '-'],
        ['Kelgan sana:', guest.arrival_date.strftime('%d.%m.%Y') if guest.arrival_date else '-'],
        ['Ketgan sana:', guest.departure_date.strftime('%d.%m.%Y') if guest.departure_date else '-'],
        ['Tashrif maqsadi:', guest.visit_purpose or '-'],
        ['Asos raqami:', guest.reference_number or '-'],
        ['Ko\'rsatilgan xizmatlar:', guest.services_provided or '-'],
        ['Restoran xarajatlari:', f"{guest.restaurant_expense:,.0f} so'm" if guest.restaurant_expense else '0 so\'m'],
        ['Sovg\'a xarajatlari:', f"{guest.gift_expense:,.0f} so'm" if guest.gift_expense else '0 so\'m'],
        ['Boshqa xarajatlar:', f"{guest.other_expenses:,.0f} so'm" if guest.other_expenses else '0 so\'m'],
        ['Jami xarajat:', f"{guest.total_expense:,.0f} so'm" if guest.total_expense else '0 so\'m'],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8e8e8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'mehmon_{guest.full_name}_{datetime.now().strftime("%Y%m%d")}.pdf',
        mimetype='application/pdf'
    )

# ==================== CELEBRATIONS MODULE ====================

@app.route('/celebrations')
@login_required
@module_access_required('celebrations')
def celebrations():
    celebrations = Celebration.query.order_by(Celebration.date.desc()).all()
    return render_template('celebrations/index.html', celebrations=celebrations)

@app.route('/celebrations/birthdays')
@login_required
@module_access_required('celebrations')
def celebrations_birthdays():
    # Get all users with birthdays
    users = User.query.filter(User.is_active == True).all()
    
    # Get this month's birthdays
    current_month = datetime.now().month
    this_month_birthdays = []
    
    for user in users:
        if user.created_at:
            if user.created_at.month == current_month:
                this_month_birthdays.append(user)
    
    return render_template('celebrations/birthdays.html', 
                         birthdays=this_month_birthdays,
                         current_month=current_month)

@app.route('/celebrations/create', methods=['GET', 'POST'])
@login_required
@module_access_required('celebrations')
def celebrations_create():
    if request.method == 'POST':
        celebration = Celebration(
            type=request.form.get('type'),
            title=request.form.get('title'),
            recipient_id=int(request.form.get('recipient_id')) if request.form.get('recipient_id') else None,
            date=datetime.strptime(request.form.get('date'), '%Y-%m-%d') if request.form.get('date') else None,
            gift_description=request.form.get('gift_description'),
            gift_value=float(request.form.get('gift_value')) if request.form.get('gift_value') else None,
            message=request.form.get('message'),
            status=request.form.get('status'),
            created_by=current_user.id
        )
        
        db.session.add(celebration)
        db.session.commit()
        
        flash('Tabriknoma qo\'shildi', 'success')
        return redirect(url_for('celebrations'))
    
    users = User.query.filter_by(is_active=True).all()
    return render_template('celebrations/create.html', users=users)

# ==================== CONTRACTS MODULE ====================

@app.route('/contracts')
@login_required
@module_access_required('contracts')
def contracts():
    contracts = Contract.query.order_by(Contract.created_at.desc()).all()
    return render_template('contracts/index.html', contracts=contracts)

@app.route('/contracts/create', methods=['GET', 'POST'])
@login_required
@module_access_required('contracts')
def contracts_create():
    if request.method == 'POST':
        contract = Contract(
            contract_number=request.form.get('contract_number'),
            contract_date=datetime.strptime(request.form.get('contract_date'), '%Y-%m-%d') if request.form.get('contract_date') else None,
            company_name=request.form.get('company_name'),
            contract_amount=float(request.form.get('contract_amount')) if request.form.get('contract_amount') else None,
            payment_date=datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d') if request.form.get('payment_date') else None,
            status=request.form.get('status'),
            description=request.form.get('description'),
            notes=request.form.get('notes'),
            created_by=current_user.id
        )
        
        db.session.add(contract)
        db.session.commit()
        
        # Handle file uploads
        if 'documents' in request.files:
            files = request.files.getlist('documents')
            for file in files:
                if file and allowed_file(file.filename):
                    filepath = save_file(file, 'uploads/contracts')
                    if filepath:
                        doc = ContractDocument(
                            contract_id=contract.id,
                            filename=file.filename,
                            filepath=filepath,
                            document_type=file.filename.rsplit('.', 1)[1].lower()
                        )
                        db.session.add(doc)
            db.session.commit()
        
        flash('Shartnoma qo\'shildi', 'success')
        return redirect(url_for('contracts'))
    
    return render_template('contracts/create.html')

@app.route('/contracts/export-excel')
@login_required
@module_access_required('contracts')
def contracts_export_excel():
    contracts = Contract.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Shartnomalar"
    
    headers = ['№', 'Shartnoma raqami', 'Sana', 'Firma nomi', 'Summa', 
               'To\'lov sanasi', 'Holati', 'Izoh']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for idx, contract in enumerate(contracts, 1):
        contract_date = contract.contract_date.strftime('%d.%m.%Y') if contract.contract_date else '-'
        payment_date = contract.payment_date.strftime('%d.%m.%Y') if contract.payment_date else '-'
        
        ws.append([
            idx,
            contract.contract_number or '-',
            contract_date,
            contract.company_name or '-',
            contract.contract_amount or 0,
            payment_date,
            contract.status or '-',
            contract.notes or '-'
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'shartnomalar_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== ADMIN PANEL ====================

@app.route('/admin')
@login_required
@admin_required
def admin_panel():
    total_users = User.query.count()
    active_users = User.query.filter_by(is_active=True).count()
    total_tasks = Task.query.count()
    recent_activities = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(20).all()
    
    return render_template('admin/panel.html',
                         total_users=total_users,
                         active_users=active_users,
                         total_tasks=total_tasks,
                         recent_activities=recent_activities)

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/<int:id>/toggle-status', methods=['POST'])
@login_required
@admin_required
def admin_toggle_user_status(id):
    user = User.query.get_or_404(id)
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'faollashtirildi' if user.is_active else 'faolsizlantirildi'
    flash(f'Foydalanuvchi {status}', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:id>/change-role', methods=['POST'])
@login_required
@admin_required
def admin_change_user_role(id):
    user = User.query.get_or_404(id)
    new_role = request.form.get('role')
    user.role = new_role
    db.session.commit()
    
    flash('Foydalanuvchi roli o\'zgartirildi', 'success')
    return redirect(url_for('admin_users'))

# ==================== API ENDPOINTS ====================

@app.route('/api/notifications')
@login_required
def api_notifications():
    notifications = Notification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).order_by(Notification.created_at.desc()).all()
    
    return jsonify({
        'count': len(notifications),
        'notifications': [{
            'id': n.id,
            'title': n.title,
            'message': n.message,
            'type': n.type,
            'link': n.link,
            'created_at': n.created_at.strftime('%d.%m.%Y %H:%M')
        } for n in notifications]
    })

@app.route('/api/notifications/<int:id>/mark-read', methods=['POST'])
@login_required
def api_mark_notification_read(id):
    notification = Notification.query.get_or_404(id)
    if notification.user_id == current_user.id:
        notification.is_read = True
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False}), 403

@app.route('/api/tasks/stats')
@login_required
def api_tasks_stats():
    if current_user.role in ['admin', 'rahbar']:
        tasks = Task.query.all()
    else:
        user_task_ids = [ta.task_id for ta in current_user.tasks_assigned]
        tasks = Task.query.filter(Task.id.in_(user_task_ids)).all()
    
    stats = {
        'total': len(tasks),
        'pending': len([t for t in tasks if t.status == 'pending']),
        'in_progress': len([t for t in tasks if t.status == 'in_progress']),
        'review': len([t for t in tasks if t.status == 'review']),
        'completed': len([t for t in tasks if t.status == 'completed']),
        'overdue': len([t for t in tasks if t.due_date and t.due_date < datetime.utcnow() and t.status != 'completed'])
    }
    
    return jsonify(stats)

@app.route('/api/dashboard/chart-data')
@login_required
def api_dashboard_chart_data():
    # Tasks by status
    if current_user.role in ['admin', 'rahbar']:
        tasks_by_status = db.session.query(
            Task.status, 
            func.count(Task.id)
        ).group_by(Task.status).all()
    else:
        user_task_ids = [ta.task_id for ta in current_user.tasks_assigned]
        tasks_by_status = db.session.query(
            Task.status, 
            func.count(Task.id)
        ).filter(Task.id.in_(user_task_ids)).group_by(Task.status).all()
    
    # Monthly tasks creation
    monthly_tasks = db.session.query(
        func.strftime('%Y-%m', Task.created_at).label('month'),
        func.count(Task.id)
    ).group_by('month').order_by('month').limit(6).all()
    
    return jsonify({
        'tasks_by_status': [{'status': status, 'count': count} for status, count in tasks_by_status],
        'monthly_tasks': [{'month': month, 'count': count} for month, count in monthly_tasks]
    })

# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('errors/403.html'), 403

# ==================== CLI COMMANDS ====================

@app.cli.command()
def init_db():
    """Initialize the database."""
    with app.app_context():
        db.create_all()
        print('Database initialized!')

@app.cli.command()
def create_admin():
    """Create admin user."""
    with app.app_context():
        admin = User.query.filter_by(email='admin@afimperiya.uz').first()
        if not admin:
            admin = User(
                full_name='Administrator',
                email='admin@afimperiya.uz',
                role='admin',
                is_active=True
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('Admin user created!')
            print('Email: admin@afimperiya.uz')
            print('Password: admin123')
        else:
            print('Admin user already exists!')

@app.cli.command()
def seed_data():
    """Seed database with sample data."""
    with app.app_context():
        # Create rahbar user
        if not User.query.filter_by(email='rahbar@afimperiya.uz').first():
            rahbar = User(
                full_name='Rahbar Rahbarovich',
                email='rahbar@afimperiya.uz',
                role='rahbar',
                department='Boshqaruv',
                position='Rahbar',
                is_active=True
            )
            rahbar.set_password('rahbar123')
            db.session.add(rahbar)
        
        # Create xodim user
        if not User.query.filter_by(email='xodim@afimperiya.uz').first():
            xodim = User(
                full_name='Xodim Xodimov',
                email='xodim@afimperiya.uz',
                role='xodim',
                department='IT',
                position='Dasturchi',
                is_active=True
            )
            xodim.set_password('xodim123')
            db.session.add(xodim)
        
        db.session.commit()
        print('Sample data seeded!')

# ==================== TELEGRAM BOT ====================

@app.route('/telegram-webhook', methods=['POST'])
def telegram_webhook():
    """Handle Telegram bot webhooks"""
    try:
        data = request.get_json()
        
        if 'message' in data:
            message = data['message']
            chat_id = message['chat']['id']
            text = message.get('text', '')
            
            # Handle /start command
            if text.startswith('/start'):
                # Extract user email or registration code if provided
                parts = text.split()
                if len(parts) > 1:
                    email_or_code = parts[1]
                    user = User.query.filter_by(email=email_or_code).first()
                    
                    if user:
                        user.telegram_chat_id = str(chat_id)
                        db.session.commit()
                        
                        response = (
                            f"✅ Xush kelibsiz, {user.full_name}!\n\n"
                            "Telegram bot muvaffaqiyatli ulandi. "
                            "Endi siz yangi topshiriqlar va bildirishnomalarni Telegram orqali olasiz."
                        )
                    else:
                        response = (
                            "❌ Foydalanuvchi topilmadi.\n\n"
                            "Iltimos, tizimda ro'yxatdan o'tganingizdan so'ng "
                            "profil sahifasidan Telegram botni ulang."
                        )
                else:
                    response = (
                        "👋 AF IMPERIYA boshqaruv tizimiga xush kelibsiz!\n\n"
                        "Bot orqali yangi topshiriqlar va bildirishnomalarni olasiz.\n\n"
                        "Botni ulash uchun:\n"
                        "1. Tizimga kiring (af-imperiya.uz)\n"
                        "2. Profilingizga o'ting\n"
                        "3. 'Telegram botni ulash' tugmasini bosing"
                    )
                
                # Send response
                import requests
                url = f"https://api.telegram.org/bot{app.config['TELEGRAM_BOT_TOKEN']}/sendMessage"
                requests.post(url, json={
                    'chat_id': chat_id,
                    'text': response,
                    'parse_mode': 'HTML'
                })
        
        return jsonify({'ok': True})
    except Exception as e:
        print(f"Telegram webhook error: {e}")
        return jsonify({'ok': False}), 500

# ==================== MAIN ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Create admin if doesn't exist
        admin = User.query.filter_by(email='admin@afimperiya.uz').first()
        if not admin:
            admin = User(
                full_name='Administrator',
                email='admin@afimperiya.uz',
                role='admin',
                is_active=True
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('\n=== ADMIN USER CREATED ===')
            print('Email: admin@afimperiya.uz')
            print('Password: admin123')
            print('===========================\n')
    
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

